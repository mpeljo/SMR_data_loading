# populate_rockprops_loader.py
# Matti Peljo   April 2024
#
# Updated May 2024.
#   + Added ability to enter multiple Bayesian percentile water content
#       values per SMR depth
#   + Moved SMR_config.py to the Config folder

# An application to transfer processed "Surface" Nuclear Magnetic
#   Resonance (SMR) groundwater properties into Geoscience Australia's
#   Excel upload template in preparation for loading into their ROCKPROPS
#   rock properties Oracle database.
#
#  SMR data is attached to Oracle "measured sections", which must exist
#   prior to loading the rest of the data to ROCKPROPS.

# To do:
#   Consider rewriting the add_section_intervals function so that, instead
#   of passing the workbook to cascading functions (which makes it confusing
#   to follow the logic), modify the add_samples and add_scalar_properties
#   functions with the relevant workbook sheets as inputs.

import os
import math
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook

# SMR_config contains all the stuff that won't change in this data load,
#   sorted by the tab (sheet) in the Oracle database Excel loader template.
from Config.SMR_config import *


def round_up(n, decimals: int = 0):
    '''
    Rounds a number towards the right of the number line,
    to an arbitrary number of decimal places
    '''
    multiplier = 10**decimals
    return math.ceil(n * multiplier) / multiplier


def remove_hints_from_template(template_xlsx: Workbook) -> Workbook:
    '''
    Every Oracle loader template Excel sheet has header information up to
    row 5 inclusive.
    This function deletes all rows after the header in each sheet used
    to load SMR data into ROCKPROPS.
    
        Parameters:
            template_xlsx:  An Excel workbook that conforms to
              Measure_Section-Samples-downhole intervals-rockprops_2023.10.

        Returns:
            template_xlsx: The Excel template with non-header rows removed.
    '''
    worksheets = [
        "SECTION INTERVAL COLLECTION",
        "SECTION INTERVALS",
        "SAMPLES",
        "SCALAR PROPERTIES",
        #'MAGNETIC REMANENCE',      Not used for SMR loading
    ]

    for sheet in worksheets:
        ws = template_xlsx[sheet]
        for row_nr in range(ws.max_row, 5, -1):
            ws.delete_rows(row_nr)

    return template_xlsx


def create_output_loader(outfile: str, loading_template: Workbook) -> str:
    '''
    Creates an empty Excel load template file
    '''
    if os.path.exists(outfile):
        print("Deleting existing SMR ROCKPROPS loading workbook")
        os.remove(outfile)
    print("Creating an empty load template file")
    loading_template.save(outfile)
    return outfile


def get_smr_df(smr_data_path: str) -> pd.DataFrame:
    '''
    Reads a .csv file containing processed SMR data and calculates an
    uncertainty value.
    
        Parameters:
            smr_data_path: The path to a file containing processed
            smr data. The file must contain these columns:
                site_name, depth_from, depth_to, p5, p50, p95
                
        Returns:
            formatted_df: A dataframe with the input columns, with depths
            rounded to centimetre precision.
    '''
    smr_df = pd.read_csv(smr_data_path)
    smr_df.rename(columns=lmh, inplace=True)
    formatted_df = smr_df.round(
        {
            "depth_from": 2,
            "depth_to": 2,
            #"BAYESIAN_LOW": 3,
            #"BAYESIAN_MEDIAN": 3,
            #"BAYESIAN_HIGH": 3,
        }
    )
    return formatted_df


def add_section_interval_collection(
        inputs: pd.core.series.Series,
        workbook: Workbook
        ) -> tuple[Workbook, str]:
    '''
    Returns an Excel workbook after appending a row of
    SECTION INTERVAL COLLECTION data to an Excel template. 
    
        Parameters:
            inputs: A pandas Series containing Measured Section database details
                Measured section name: The name of the SMR site
                ENO: The measured section's "Entity number"
                Dep ref point: The primary key of measured section's
                    "depth reference point", usually ground level
                Acq_date: The date the SMR dataset was collected in the field
            workbook: A clean Excel loading template conforming to
                Measure_Section-Samples-downhole intervals-rockprops_2023.10 
                
        Returns:
            workbook: An Excel workbook with a row of data appended to the
                SECTION INTERVAL COLLECTION worksheet
            site_name: A string containing the current site name
    '''
    # Inputs are: Measured section name, ENO, Dep ref point
    site_name = inputs["Measured section name"]
    print(f"Adding SECTION INTERVAL COLLECTION for: {site_name}")

    ws = workbook["SECTION INTERVAL COLLECTION"]
    collection_name = site_name + COLLECTION_NAME_POSTFIX
    new_row = [
        inputs["ENO"],                      # Section ENO
        inputs["Measured section name"],    # Measured section name
        inputs["Dep ref point"],            # Depth reference point ID
        "",                                 # Interval collection id
        collection_name,                    # Collection name
        COLLECTION_TYPE,                    # Collection type
        ORIGINATOR_NUMBER,                  # Originator number
        PREFERRED_COLLECTION,               # Preferred collection
        "",                                 # Collection source document id
        "",                                 # Source comments
    ]
    #print(new_row)

    ws.append(new_row)
    return workbook, site_name


def add_section_intervals(
        working_info: pd.core.series.Series,
        site_name: str,
        smr_water_data_path: str,
        workbook: Workbook
        ) -> Workbook:
    '''
    Returns a completed Excel workbook after appending SMR data in the
        * SECTION INTERVALS,
        * SAMPLES, and
        * SCALAR PROPERTIES Excel template worksheets.

        Calls the "add_sample" function, which calls the "add_scalar_properties"
        function to populate the SAMPLES and SCALAR PROPERTIES sheets,
        cascading certain variables from one sheet to the next. 
    
        Parameters:
            working_info: A pandas Series containing Measured Section database details
                Measured section name: The name of the SMR site
                ENO: The measured section's "Entity number"
                Dep ref point: The primary key of measured section's
                    "depth reference point", usually ground level
                Acq_date: The date the SMR dataset was collected in the field
            site_name: A string containing the current site name
            smr_water_data_path: A string containing a path to the processed
                SMR data
            workbook: The Excel workbook containing the row created by the
                add_section_interval_collection function

        Returns:
            workbook: An Excel workbook containing completed SMR worksheets

    '''
    print(f"Adding SECTION INTERVAL for: {site_name}")
    intervals = workbook["SECTION INTERVALS"]
    collection_name = site_name + COLLECTION_NAME_POSTFIX
    smr_data = get_smr_df(smr_water_data_path)
    # smr_data contains:
    #   site_name, depth_from, depth_to, p5, p50, p95

    # Get the subset of results that is relevant to the current site
    current_results = smr_data[(smr_data["site_name"] == site_name)]
    current_results.reset_index(inplace=True)
    print(f'Adding samples and scalar properties information for collection: {collection_name}')
    for idx, row in current_results.iterrows():
        depth_from = int(row["depth_from"])
        depth_to = int(row["depth_to"])
        if depth_to > MAX_DEPTH:
            break
        interval_id = f"{site_name}_{depth_from}-{depth_to}m"
        new_row = [
            "",                         # Collection no.
            collection_name,            # Collection name
            "",                         # Interval no.
            interval_id,                # Interval ID
            row["depth_from"],          # Interval start depth
            row["depth_to"],            # Interval end depth
            INTERVAL_UNIT_OF_MEASURE,   # Interval unit of measure
        ]
        print(new_row)
        intervals.append(new_row)
        workbook = add_samples(
            working_info, collection_name, interval_id, workbook, idx, row
        )
    return workbook


def add_samples(
        working_info: pd.core.series.Series,
        collection_name: str,
        interval_id: str,
        workbook: Workbook,
        smr_results_idx,
        smr_results_row: pd.core.series.Series,
        ) -> Workbook:
    '''
    Returns a completed Excel workbook after appending SMR data in the
        * SAMPLES, and
        * SCALAR PROPERTIES Excel template worksheets.

        Calls the "add_scalar_properties" function to populate the
        SCALAR PROPERTIES sheet, cascading certain variables to the next sheet. 
    
        Parameters:
            working_info: A pandas Series containing Measured Section database details
                Measured section name: The name of the SMR site
                ENO: The measured section's "Entity number"
                Dep ref point: The primary key of measured section's
                    "depth reference point", usually ground level
                Acq_date: The date the SMR dataset was collected in the field
            collection_name: A string containing the current collection name
            workbook: The Excel workbook containing the row created by the
                add_section_interval_collection function
            smr_results_idx: The 0-based index value of the current collection
                of SMR results
            smr_results_row: A Series comprising the processed SMR results that
                are relevant to the current sample. The Series contains:
                site_name, depth_from, depth_to, p5, p50, p95, uncert

        Returns:
            workbook: An Excel workbook containing completed SMR worksheets
    '''
    # Working_info contains:
    # Measured section name, ENO, Dep ref point, Acq_date
    samples = workbook["SAMPLES"]
    sample_id = f'{smr_results_row["site_name"]}_L{smr_results_idx+1}'
    new_row = [
        working_info["ENO"],        # Section eno
        collection_name,            # Collection name
        "",                         # Intervalno
        "",                         # Sampleno
        interval_id,                # Interval id
        sample_id,                  # Sample id
        working_info["Acq_date"],   # Acquisition date
        "",                         # Parent sampleno
        "",                         # Parent sample id
        ANO,                        # ANO
        ACCESS_CODE,                # Access code
        "",                         # Confidential until date
        QA_STATUS,                  # QA status
        ACTIVITY_CODE,              # Activity code
        SAMPLE_TYPE,                # Sample type
        SAMPLING_METHOD,            # Sampling method
        MATERIAL_CLASS,             # Material class
        "",                         # Procedure no
        PROJECT_NO,                 # Project no
        "",                         # Refid
        "",                         # Other id
        "",                         # Specimen storage location
        "",                         # Storage date
        "",                         # Comments about specimen or observation
        "",                         # ISGN
        "",                         # Specimen mass
        "",                         # Mass UOM
        "",                         # Source
    ]
    #print(new_row)
    samples.append(new_row)
    workbook = add_scalar_properties(workbook, sample_id, smr_results_row)
    return workbook


def add_scalar_properties(
        workbook: Workbook,
        sample_id: str,
        smr_results_row: pd.core.series.Series) -> Workbook:
    '''
    Returns a completed Excel workbook after appending SMR data in the
        SCALAR PROPERTIES Excel template worksheets.
    
        Parameters:
            workbook: The Excel workbook containing the row created by the
                add_section_interval_collection function
            sample_id: The ID of the sample that was created by the
                add_section_intervals function
            smr_results_row: A Series comprising the processed SMR results that
                are relevant to the current sample. The Series contains:
                site_name, depth_from, depth_to, p5, p50, p95, uncert

        Returns:
            workbook: An Excel workbook containing completed SMR worksheets
    '''
    scalar_properties = workbook["SCALAR PROPERTIES"]

    for _, (_, bayesian_percentile) in enumerate(lmh.items()):
        print(bayesian_percentile)
        current_bayesian_constants = BAYESIAN_CONSTANTS[bayesian_percentile]
        measurement = round_up(smr_results_row[bayesian_percentile], 4)
        if np.isnan(measurement):
            measurement = NULL_VALUE

        new_row = [
            sample_id,                  # SAMPLEID
            "",                         # Sample number
            "",                         # Resultno
            "",                         # Resultid
            ACCESS_CODE,                # Access code
            "",                         # Confidential until date
            QA_STATUS,                  # QA status
            ORIGNO,                     # Originator
            SOURCE_TYPE,                # Source type
            SOURCE,                     # Source
            LOAD_APPROVED,              # Load approved
            PROCESS_TYPE,               # Process type (PROCESS.PROCESSNO)
            PETROPHYSICAL_PROPERTY,     # Petrophysical property
            measurement,                # Value
            UOM,                        # UOM
            current_bayesian_constants["RESULT QUALIFIER"],     # Result qualifier
            current_bayesian_constants["UNCERTAINTY TYPE"],     # uncertainty type
            current_bayesian_constants["UNCERTAINTY VALUE"],    # UNCERTAINTY VALUE
            current_bayesian_constants["UNCERTAINTY UOM"],      # UNCERTAINTY UOM
            "",                         # RESULT DATE/TIME
            "",                         # OBSERVATION LOCATION
            current_bayesian_constants['SP_REMARKS'],                 # REMARKS
            NUMERICAL_CONFIDENCE,       # Numerical confidence
            METADATA_QUALITY,           # Metadata quality
            SUMMARY_CONFIDENCE,         # Summary confidence
        ]

        #print(new_row)
        scalar_properties.append(new_row)
        print(f"Scalar properties current row count: {scalar_properties.max_row}")

    return workbook


def main(development_stage="test"):
    # Set up input and output file locations based on development stage of
    #   this database product.
    if development_stage == "final":
        template_file = full_template_file
        output_file = final_output_file
    elif development_stage == "test":
        template_file = full_template_file
        output_file = test_output_file

    # Remove rows with hints from template Excel workbook
    full_wb = load_workbook(filename=template_file)
    clean_template_wb = remove_hints_from_template(full_wb)

    # Create output workbook
    output_path = output_dir + output_file
    output_loader = create_output_loader(output_path, clean_template_wb)
    wb = load_workbook(filename=output_loader)

    # Read section name, eno, depth reference point number, acquisition date
    smr_metadata = pd.read_excel(config_file)
    pd.to_datetime(smr_metadata["Acq_date"], dayfirst=True)

    # Populate the Excel template
    for row_nr in range(smr_metadata.shape[0]):
        working_info = smr_metadata.iloc[row_nr]
        # Working_info contains:
        # Measured section name, ENO, Dep ref point, Acq_date
        print(f"\nProcessing data for site: {working_info['Measured section name']}")
        wb, site_name = add_section_interval_collection(working_info, wb)
        wb = add_section_intervals(working_info, site_name, smr_water_data_path, wb)

    wb.save(output_loader)


if __name__ == "__main__":
    # Set up input and output file locations based on development stage of
    #   this database product.

    main(STAGE)
